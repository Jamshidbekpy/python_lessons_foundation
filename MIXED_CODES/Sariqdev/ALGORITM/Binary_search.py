from openpyxl import Workbook, load_workbook
wb = Workbook()
ws = wb.active

workbook = load_workbook('binary_search.xlsx')
sheet = workbook['data']

# Binary search
search = input()
list_inf = []
for row in range(1, sheet.max_row+1):
    list_i = []
    for column in range(1, sheet.max_column+1):
        cell_value = sheet.cell(row=row, column=column).value
        list_i.append(cell_value)
    list_inf.append(list_i)
print(list_inf)
