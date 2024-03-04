from openpyxl import load_workbook, Workbook
# path = "C:\WEBDASTURLASH\EXEL\Exel-tajriba.xlsx"
# workbook = load_workbook(path)
# sheet = workbook['Sheet1']   # Sheet1 sheet ga yuklandi buni o'zi bitta lug'at
#
# sheet['A1'] = 4
# from openpyxl import load_workbook

# Excel faylini yuklash
workbook = load_workbook("C:\WEBDASTURLASH\EXEL\Exel-tajriba.xlsx")

# Ma'lum bir jadvalni olish
sheet = workbook['Sheet1']

# Ma'lum bir o'rni o'qish va yozish
value = sheet['A1'].value
for i in range(6):
    sheet[f'B{str(i+1)}'] = 4
    print(sheet[f'A{str(i+1)}'].value)
    print(sheet[f'B{str(i + 1)}'].value)

