from openpyxl import load_workbook
from datetime import datetime
import os

def product_bolim_name()->list:
    list_bolim = ['','Mevalar','Ichimliklar']
    for index,bolim in enumerate(list_bolim):
        if index == 0:
            continue
        print(f'{index} --> {bolim}')
    select = input('Bolimni tanlang/0: ')
    if select.isdigit() and int(select) > 0 and int(select) < len(list_bolim):
        select = list_bolim[int(select)]
        select_name = input('Mahsulot nomini kiriting:')
        return [select,select_name]
    elif select == '0':
        main()        
    else :
        print('Bo\'lim nomi xato kiritildi!Iltimos boshqatdan kiriting!')
        main()
def product_miqdori()->int:
    while True:
        miqdor = input('Miqdorni kiriting:')
        if miqdor.isdigit() and int(miqdor) > 0:
            return int(miqdor)
        else:
            print('Miqdor xato kiritildi!')
def product_narxi()->int:
    while True:
        narx = input('Narxni kiriting:')
        if narx.isdigit() and int(narx) > 0:
            return int(narx)
        else:
            print('Narx xato kiritildi!')
       

def chiqarilgan_sana()->datetime:
    while True:
        chiqarilgan_sana = input('Chiqarilgan sana kiriting *kun/oy/yil*:')
        if  chiqarilgan_sana.count('/') == 2 and datetime.strptime(chiqarilgan_sana,'%d/%m/%Y'):
            if datetime.strptime(chiqarilgan_sana,'%d/%m/%Y') < datetime.now():
                return datetime.strptime(chiqarilgan_sana,'%d/%m/%Y')
        else:
            print('Chiqarilgan sana xato kiritildi!')  
def product_muddat()->datetime:
    while True:
        muddat = input('Muddat kiriting: *kun/oy/yil*:')
        if  muddat.count('/') == 2 and datetime.strptime(muddat,'%d/%m/%Y'):
            if datetime.strptime(muddat,'%d/%m/%Y') > datetime.now():
                return datetime.strptime(muddat,'%d/%m/%Y')
        else:
            print('Muddat xato kiritildi!')
def product_country()->str:
    country = input('Ishlab chiqarilgan mamlakatini kiriting:')
    return country   
# ------------------------------------------------------------------------------------------

def add_product()->None:
    l_bolim_name = product_bolim_name() # [select,select_name]
    print(l_bolim_name)
    database = load_workbook('DATABASE.xlsx')
    sheet = database[l_bolim_name[0]]
    def tekshir():
        for i in range(2,sheet.max_row+1):
            if sheet[f'A{i}'].value == l_bolim_name[1]:
                return i
        return False
    if tekshir() !=False :
        print(sheet[f'B{tekshir()}'].value)
        miqdor = product_miqdori()
        narx = sheet[f'C{tekshir()}'].value
        kelgan_sanasi = datetime.now()
        sheet[f'B{tekshir()}'] = sheet[f'B{tekshir()}'].value + miqdor
        database.save('DATABASE.xlsx')
        with open('add_report.txt', 'a') as file:
            file.write(f'Mahsulot nomi: {l_bolim_name[1]}\n')
            file.write(f'Miqdor: {miqdor}\n')
            file.write(f'Narx: {narx}\n')
            file.write(f'Kelgan sanasi: {kelgan_sanasi}\n')
            file.write(f'Jami summa: {miqdor * narx}\n') 
            file.write('\n')
            file.close()
    else:
        miqdor = product_miqdori()
        narx = product_narxi()
        kelgan_sanasi = datetime.now()
        chiqarilgan = chiqarilgan_sana()
        muddat = product_muddat()
        country = product_country()
        sheet.append([l_bolim_name[1],miqdor,narx,chiqarilgan,muddat,country])
        database.save('DATABASE.xlsx')
        with open('add_report.txt', 'a') as file:
            file.write(f'Mahsulot nomi: {l_bolim_name[1]}\n')
            file.write(f'Miqdor: {miqdor}\n')
            file.write(f'Narx: {narx}\n')
            file.write(f'Kelgan sanasi: {kelgan_sanasi}\n')
            file.write(f'Jami summa: {miqdor * narx}\n') 
            file.write('\n')
            file.close()
    print('Mahsulot muvaffaqiyatli qo\'shildi!')
    main()         
def sell_product()->None:
    l_bolim_name = product_bolim_name() # [select,select_name]
    database = load_workbook('DATABASE.xlsx')
    sheet = database[l_bolim_name[0]]
    def tekshir():
        for i in range(2,sheet.max_row+1):
            if sheet[f'A{i}'].value == l_bolim_name[1]:
                return i
        return False
    if tekshir() != False:
        print(sheet[f'B{tekshir()}'].value)
        miqdor = product_miqdori()
        narx = sheet[f'C{tekshir()}'].value
        sotilgan_sanasi = datetime.now()
        mavjud = sheet[f'B{tekshir()}'].value
        if sheet[f'B{tekshir()}'].value >= miqdor:
            sheet[f'B{tekshir()}'] = sheet[f'B{tekshir()}'].value - miqdor
            database.save('DATABASE.xlsx')
            print(f'Siz {miqdor} miqdorda {l_bolim_name[1]} mahsulotidan sotdingiz!')
            with open('sell_report.txt', 'a') as file:
                file.write(f'Mahsulot nomi: {l_bolim_name[1]}\n')
                file.write(f'Miqdor: {miqdor}\n')
                file.write(f'Narx: {narx}\n')
                file.write(f'Kelgan sanasi: {sotilgan_sanasi}\n')
                file.write(f'Jami summa: {(miqdor * narx)/100*105}\n') # 5% komissiyasi
                file.write('\n')
                file.close()
            main()
        else :
            print(f'Siz sotadigan mahsulot bizda {mavjud} miqdorda qolgan!')
            sell_product()
    else:
        print('Siz kiritgan mahsulot bizda mavjud emas!')
        sell_product()
def report():
    select  = input('Qanday hisobotni ko\'rmoqchisiz? 1->Add/2->Sell/3->Ombor/0->Exit:')
    if select == '0':
        main()
    elif select == '1':
        os.system('add_report.txt')
        report()
    elif select == '2':
        os.system('sell_report.txt')
        report()
    elif select == '3':
        os.startfile('DATABASE.xlsx')
        report()
    else:
        print('Xato tanlov!Boshqatdan tanlang!')
        report() 
def remove(): 
    if os.path.exists('add_report.txt'):
        os.remove('add_report.txt') 
    if os.path.exists('sell_report.txt'):
        os.remove('sell_report.txt')
    exit()  
    
def main():
    select = input('1. Add\n2. Sell\n3. Report\n0. Exit\nTanlang:')
    if select == '1':
        add_product()
    elif select == '2':
        sell_product()
    elif select == '3':
        report()
    elif select == '0':
       remove() 
    else:
        print('Xato kiritildi')
        main()
main()


