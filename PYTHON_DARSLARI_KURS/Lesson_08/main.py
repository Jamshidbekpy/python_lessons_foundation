from openpyxl import Workbook,load_workbook
from datetime import datetime
from os import rmdir, path
wb = Workbook()
ws = wb.active

workbook = load_workbook('MOCK_DATA.xlsx')
sheet = workbook['data']

list_question = ["",
                 "1)Ushbu jadvaldan oyligi 1000 $ dan yuqori bo'lgan ishchilarni saralab chiqaring",
                 "2)Jadvaldan ism yoki familiya bo'yicha turli insonlarni qidiring.Jadvalda mavjud bo'lsa,u haqidai barcha ma'lumotlarni chiroyli ko'rinishda qaytaring.Bo'lmasa,ushbu inson bizning ro'yxatda mavjud emas yozuvini chiqaring.",
                 "3)Hozirda 50 yoshdan katta shaxslarni saralang.Ularning barcha ma'lumotlarni consolega chiqaring.",
                 "4)Yangi sheet qo'shing.Bu sheetga 25 yoshli ishchilar ro'yxatini chiqaring.Ro'yxatda ularning barch ma'lumotlari aniq holda qaytarilsin.",
                 "5)Exelda berilgan ushbu ma'lumotlarni dictonary ko'rnishida matnli faylga o'tkazing.",
                 "6)Matnli fayldagi ma'lumotdan 30 yoshli shaxslarni ma'lumotlarini yangi exel faylga saqlang.",
                 "7)Jadvaldan shaxslarni first_name,last_name,salery si bo'yicha yangi sheetga nusxa ko'chiring.",
                 "8)#Jurnaldagi shaxslarni oyligi bo'yicha diagramma qiling."
                 ]


while True:
    for i in list_question:
        print(i, "\n")
    number = input("Savol nomeri:")

    if number.isdigit() and int(number) >0 and int(number) <=9:
        if int(number) == 1:
            list_info = []
            for row in range(2, sheet.max_row + 1):
                list_i = []
                for column in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=column).value
                    list_i.append(cell_value)
                list_info.append(list_i)
            print(list_info)


        elif int(number) == 2:
            search_fl_name = input('Ism yoki familiya kiriting:')
            n = 0
            for row in range(2, sheet.max_row + 1):
                string_f_name = str(sheet[f'A{row}'].value)
                string_l_name = str(sheet[f'B{row}'].value)
                if search_fl_name == (string_f_name or string_l_name):
                    n += 1
                    print(sheet[f'A1'].value, ":", string_f_name, "    ", sheet[f'B1'].value, ":", string_l_name, "    ", sheet[f'C1'].value, ":", sheet[f'C{row}'].value, "    ",sheet[f'D1'].value, ":", sheet[f'D{row}'].value, "    ", sheet[f'E1'].value, ":", sheet[f'E{row}'].value, "    ", sheet[f'F1'].value, ":", sheet[f'F{row}'].value, "    ", sheet[f'G1'].value, ":", sheet[f'G{row}'].value, "    ", sheet[f'H1'].value, ":", sheet[f'H{row}'].value, "    ", sheet[f'I1'].value, ":", sheet[f'I{row}'].value)
            if n == 0:
                print("Bunday ismli shaxs topilmadi!")

        elif int(number) == 3:
           pass
        elif int(number) == 4:
           pass

        elif int(number) == 5:
            pass
        elif int(number) == 6:

            pass
        elif int(number) == 7:
            pass
        elif int(number) == 8:
            pass
        elif int(number) == 9:
            select = input("O'chirish fayl nomini to'liq kiriting:")
            if path.exists(select):
                rmdir(select)

            else:
                print("Bunday fayl yo'q!")
    else:
        print("Savol nomerini xato kiritdingiz!")