from main_abc import ShopStructure
from product import Product
from openpyxl import load_workbook
from os import path,startfile,remove,system
    
class Shop(ShopStructure):
    def __init__(self,name,location,data):
        self.name = name
        self.location = location
        self.data = data
        self.database = load_workbook(data)
    def add(self):
        department_name = Product.get_department_and_name() # [select,select_name]
        if department_name == 0:
            self.main()
        else:
            print(department_name)
            sheet = self.database[department_name[0]]
            def check():
                for i in range(2,sheet.max_row+1):
                    if sheet[f'A{i}'].value == department_name[1]:
                        return i
                return False
            if check() !=False :
                print(sheet[f'B{check()}'].value)
                quantity = Product.get_quantity()
                price = sheet[f'C{check()}'].value
                add_date = Product.get_date_of_arrival()
                sheet[f'B{check()}'] = sheet[f'B{check()}'].value + quantity
                self.database.save(self.data)
                with open('add_report.txt', 'a') as file:
                    file.write(f'Mahsulot nomi: {department_name[1]}\n')
                    file.write(f'Miqdor: {quantity}\n')
                    file.write(f'Narx: {price}\n')
                    file.write(f'Kelgan sanasi: {add_date}\n')
                    file.write(f'Jami summa: {quantity * price}\n') 
                    file.write('\n')
                    file.close()
            else:
                quantity = Product.get_quantity()
                price = Product.get_price()
                add_date = Product.get_date_of_arrival()
                old_date = Product.get_old_date()
                expiration = Product.get_expiration_date()
                country = Product.get_country()
                sheet.append([department_name[1],quantity,price,add_date,old_date,expiration,country])
                self.database.save(self.data)
                with open('add_report.txt', 'a') as file:
                    file.write(f'Mahsulot nomi: {department_name[1]}\n')
                    file.write(f'Miqdor: {quantity}\n')
                    file.write(f'Narx: {price}\n')
                    file.write(f'Kelgan sanasi: {add_date}\n')
                    file.write(f'Jami summa: {quantity * price}\n') 
                    file.write('\n')
                    file.close()
            print('Mahsulot muvaffaqiyatli qo\'shildi!')
            self.main()         

    def sell(self):
        department_name = Product.get_department_and_name() # [select,select_name]
        if department_name == 0:
            self.main()
        else:
            print(department_name)
            sheet = self.database[department_name[0]]
            def check():
                for i in range(2,sheet.max_row+1):
                    if sheet[f'A{i}'].value == department_name[1]:
                        return i
                return False
            if check() != False:
                print(sheet[f'B{check()}'].value)
                quantity = Product.get_quantity()
                price = sheet[f'C{check()}'].value
                sell_date = Product.get_date_of_arrival()
                have = sheet[f'B{check()}'].value
                if sheet[f'B{check()}'].value >= quantity:
                    sheet[f'B{check()}'] = sheet[f'B{check()}'].value - quantity
                    
                    self.database.save(self.data)
                    
                    print(f'Siz {quantity} miqdorda {department_name[1]} mahsulotidan sotdingiz!')
                    with open('sell_report.txt', 'a') as file:
                        file.write(f'Mahsulot nomi: {department_name[1]}\n')
                        file.write(f'Miqdor: {quantity}\n')
                        file.write(f'Narx: {price}\n')
                        file.write(f'Sotilgan sanasi: {sell_date}\n')
                        file.write(f'Jami summa: {(quantity * price)/100*105}\n') # 5% komissiyasi
                        file.write('\n')
                        file.close()
                    self.main()
                else :
                    print(f'Siz sotadigan mahsulot bizda {have} miqdorda qolgan!')
                    self.sell()
            else:
                print('Siz kiritgan mahsulot bizda mavjud emas!')
                self.sell()


    def report(self):
        select = input('Hisobotni tanlang: 1 - Add, 2 - Sell, 3 - Database, 0 - Exit:')
        if select == '0':
            self.main()
        elif select == '1':
            if path.exists('add_report.txt'):
                system('add_report.txt')
            else:
                print('Hali bazaga yangi mahsulot qo\'shilmadi!')
            self.report()
        elif select == '2':
            if path.exists('sell_report.txt'):
                system('sell_report.txt')
            else:
                print('Hali yangi mahsulot sotilmadi!')
            self.report()
        elif select == '3':
                startfile(self.data)

    def main(self):
        select = input('Bo\'limni tanlang: 1 - Add, 2 - Sell, 3 - Report,4 - Info_shop , 0 - Exit:')
        if select == '1':
            self.add()
        elif select == '2':
            self.sell()
        elif select == '3':
            self.report()
        elif select == '4':
            self.__str__()
        elif select == '0':
            if path.exists('add_report.txt'):
                remove('add_report.txt')
            if path.exists('sell_report.txt'):
                remove('sell_report.txt')
        else:
            print("Iltimos, bo'limga mos sonni tanlang!")
            self.main()

    def __str__(self):
        print(f"Ushbu do'kon nomi {self.name}. Joylashuvi {self.location}.Ma'lumotlar ba'zasi nomi {self.data}.")
        self.main()

        