from main_abc import ShopStructure
from product import Product
from openpyxl import load_workbook,Workbook
from os import path,startfile,remove,system
    
class Shop(ShopStructure):
    def __init__(self,name,location,data):
        self.name = name
        self.location = location
        self.data = data
    def add(self):
        l_department = ['']
        select = ''
        select_name = ''
        if path.exists(self.data):
            database = load_workbook(self.data)
            sheetnames = database.sheetnames
            for sheetname in sheetnames:
                l_department.append(sheetname)
            print(l_department)
            for index,department in enumerate(l_department):
                if index == 0:
                    continue
                print(f'{index} --> {department}')
            select = input('Bolimni tanlang/0: ')
            if select == '0':
                self.main()
            if select.isdigit() and int(select) > 0 and int(select) < len(l_department):
                select = l_department[int(select)]
                select_name = input('Mahsulot nomini kiriting:')
            elif select == '0':
                self.main()        
            else :
                print('Bo\'lim nomi xato kiritildi!Iltimos boshqatdan kiriting!')
                self.add()
                
        elif not path.exists(self.data):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = input('Bo\'lim nomini kiriting:')
            sheet.append(['Mahsulot nomi','Miqdori','Narxi','Kelgan sanasi','Chiqarilgan sanasi','Muddati tugaygan sana','Davlat'])
            workbook.save(self.data)
            while True:
                request = input('Yana bo\'lim qo\'shmoqchimisiz? (y/n):')
                if request == 'y':
                    sheet = workbook.create_sheet(input('Yangi bo\'lim nomini kiriting:'))
                    sheet.append(['Mahsulot nomi','Miqdori','Narxi','Kelgan sanasi','Chiqarilgan sanasi','Muddati tugaygan sana','Davlat'])
                    workbook.save(self.data)
                elif request == 'n':
                    break
                else:
                    print('Iltimos, "y" yoki "n" tugmasini bosing!')
            database = load_workbook(self.data)
            sheetnames = database.sheetnames
            for sheetname in sheetnames:
                l_department.append(sheetname)
            print(l_department)
            for index,department in enumerate(l_department):
                if index == 0:
                    continue
                print(f'{index} --> {department}')
            select = input('Bolimni tanlang/0: ')
            if select.isdigit() and int(select) > 0 and int(select) < len(l_department):
                select = l_department[int(select)]
                select_name = input('Mahsulot nomini kiriting:')
            elif select == '0':
                self.main()        
            else :
                print('Bo\'lim nomi xato kiritildi!Iltimos boshqatdan kiriting!')
                self.add()
        department_name = [select,select_name] # [select,select_name]
        if department_name == '0':
            self.main()
        elif department_name == ['0','']:
            name = input('Do\'kon nomini kiriting:')
            if name == '0':
                exit()
            location = input('Joylashuvni kiriting:')
            data = f'{name}.xlsx'
            obj = Shop(name,location,data)
            obj.main()
        else:
            print(department_name)
            sheet = database[department_name[0]]
            def check():
                for i in range(2,sheet.max_row+1):
                    if sheet[f'A{i}'].value == department_name[1]:
                        return i
                return False
            if check() !=False :
                print(sheet[f'B{check()}'].value)
                quantity = Product.get_quantity()
                price = sheet[f'C{check()}'].value
                add_date = Product.get_date_of_arrival()
                sheet[f'B{check()}'] = sheet[f'B{check()}'].value + quantity
                database.save(self.data)
                with open(f'add_report{self.name}.txt', 'a') as file:
                    file.write(f'Mahsulot nomi: {department_name[1]}\n')
                    file.write(f'Miqdor: {quantity}\n')
                    file.write(f'Narx: {price}\n')
                    file.write(f'Kelgan sanasi: {add_date}\n')
                    file.write(f'Jami summa: {quantity * price}\n') 
                    file.write('\n')
                    file.close()
            else:
                quantity = Product.get_quantity()
                price = Product.get_price()
                add_date = Product.get_date_of_arrival()
                old_date = Product.get_old_date()
                expiration = Product.get_expiration_date()
                country = Product.get_country()
                sheet.append([department_name[1],quantity,price,add_date,old_date,expiration,country])
                database.save(self.data)
                with open(f'add_report{self.name}.txt', 'a') as file:
                    file.write(f'Mahsulot nomi: {department_name[1]}\n')
                    file.write(f'Miqdor: {quantity}\n')
                    file.write(f'Narx: {price}\n')
                    file.write(f'Kelgan sanasi: {add_date}\n')
                    file.write(f'Jami summa: {quantity * price}\n') 
                    file.write('\n')
                    file.close()
            print('Mahsulot muvaffaqiyatli qo\'shildi!')
            self.main()         
   
    def sell(self):
        l_department = ['']
        select = ''
        select_name = ''
        database = None
        if path.exists(self.data):
            database = load_workbook(self.data)
            sheetnames = database.sheetnames
            for sheetname in sheetnames:
                l_department.append(sheetname)
            print(l_department)
            for index,department in enumerate(l_department):
                if index == 0:
                    continue
                print(f'{index} --> {department}')
            select = input('Bolimni tanlang/0: ')
            if select == '0':
                self.main()
            if select.isdigit() and int(select) > 0 and int(select) < len(l_department):
                select = l_department[int(select)]
                select_name = input('Mahsulot nomini kiriting:')
            elif select == '0':
                self.main()        
            else :
                print('Bo\'lim nomi xato kiritildi!Iltimos boshqatdan kiriting!')
                self.add()
                
        else:
            print('Bunday do\'kon mavjud emas!')
        department_name = [select,select_name] # [select,select_name]
        if department_name == '0':
            self.main()
        else:
            print(department_name)
            if department_name == ['0','']:
                name = input('Do\'kon nomini kiriting:')
                if name == '0':
                    exit()
                location = input('Joylashuvni kiriting:')
                data = f'{name}.xlsx'
                obj = Shop(name,location,data)
                obj.main()
            elif department_name == ['','']:
                print('Hali do\'kon mavjud emas!')
                name = input('Do\'kon nomini kiriting:')
                if name == '0':
                    exit()
                location = input('Joylashuvni kiriting:')
                data = f'{name}.xlsx'
                obj = Shop(name,location,data)
                obj.main()
            try:
                sheet = database[department_name[0]]
            except:
                name = input('Do\'kon nomini kiriting:')
                if name == '0':
                    exit()
                location = input('Joylashuvni kiriting:')
                data = f'{name}.xlsx'
                obj = Shop(name,location,data)
                obj.main()
            def check():
                for i in range(2,sheet.max_row+1):
                    if sheet[f'A{i}'].value == department_name[1]:
                        return i
                return False
            if check() != False:
                print(sheet[f'B{check()}'].value)
                quantity = Product.get_quantity()
                price = sheet[f'C{check()}'].value
                sell_date = Product.get_date_of_arrival()
                have = sheet[f'B{check()}'].value
                if sheet[f'B{check()}'].value >= quantity:
                    sheet[f'B{check()}'] = sheet[f'B{check()}'].value - quantity
                    
                    database.save(self.data)
                    
                    print(f'Siz {quantity} miqdorda {department_name[1]} mahsulotidan sotdingiz!')
                    with open(f'sell_report{self.name}.txt', 'a') as file:
                        file.write(f'Mahsulot nomi: {department_name[1]}\n')
                        file.write(f'Miqdor: {quantity}\n')
                        file.write(f'Narx: {price}\n')
                        file.write(f'Sotilgan sanasi: {sell_date}\n')
                        file.write(f'Jami summa: {(quantity * price)/100*105}\n') # 5% komissiyasi
                        file.write('\n')
                        file.close()
                    self.main()
                else :
                    print(f'Siz sotadigan mahsulot bizda {have} miqdorda qolgan!')
                    self.sell()
            else:
                print('Siz kiritgan mahsulot bizda mavjud emas!')
                self.sell()

    def report(self):
        select = input('Hisobotni tanlang: 1 - Add, 2 - Sell, 3 - Database, 0 - Exit:')
        if select == '0':
            self.main()
        elif select == '1':
            if path.exists(f'add_report{self.name}.txt'):
                system(f'add_report{self.name}.txt')
            else:
                print('Hali bazaga yangi mahsulot qo\'shilmadi!')
            self.report()
        elif select == '2':
            if path.exists(f'sell_report{self.name}.txt'):
                system(f'sell_report{self.name}.txt')
            else:
                print('Hali yangi mahsulot sotilmadi!')
            self.report()
        elif select == '3':
            if path.exists(self.data):
                startfile(self.data)
            else:
                print('Do\'konda hali ba\'za mavjud emas!')
                

    def main(self):
        select = input('Bo\'limni tanlang: 1 - Add, 2 - Sell, 3 - Report,4 - Info_shop , 0 - Exit:')
        if select == '0':
            if path.exists(f'add_report{self.name}.txt'):
                remove(f'add_report{self.name}.txt')
            if path.exists(f'sell_report{self.name}.txt'):
                remove(f'sell_report{self.name}.txt')
        elif select == '1':
            self.add()
        elif select == '2':
            self.sell()
        elif select == '3':
            self.report()
        elif select == '4':
            self.get_info()
       
        else:
            print("Iltimos, bo'limga mos sonni tanlang!")
            self.main()
    def get_info(self):
        print(f"Ushbu do'kon nomi {self.name}. Joylashuvi {self.location}.Ma'lumotlar ba'zasi nomi {self.data}.")
        self.main()

        