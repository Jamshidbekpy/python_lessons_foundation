from shop import Shop
from os import remove
from openpyxl import Workbook
def shop_create():
    name = input("Do'kon nomini kiriting:")
    location = input("Do'kon joylashuvi:")
    while True:
        data = input("Ma'lumotlar bazasi nomini kiriting:")
        try:
            return Shop(name,location,data)
        except:
            try:
                workbook = Workbook()
                sheet1 = workbook.create_sheet('Ichimliklar')
                sheet2 = workbook.create_sheet('Oziq-ovqat')
                sheet1.append(['Mahsulot nomi','Miqdori','Narxi','Birinchi kelgan mahsulot sanasi','Muddati tugaydigan sanasi','Chiqarilgan sanasi','Davlat'])
                sheet2.append(['Mahsulot nomi','Miqdori','Narxi','Birinchi kelgan mahsulot sanasi','Muddati tugaydigan sanasi','Chiqarilgan sanasi','Davlat'])
                workbook.save(data)
                return Shop(name,location,data)
            except:
                print('Excel fayl nomini nomini xato kiritdingiz!')
                remove(data)
                continue
  
while True:
    request = input("Biror do'kon bilan ishlamoqchimisiz? (y/n):")
    if request == 'y':
        shop_create().main()
    elif request == 'n':
            exit()
    else:
        print("Iltimos 'y' yoki 'n' kiritishni unutmang!")
        
   

