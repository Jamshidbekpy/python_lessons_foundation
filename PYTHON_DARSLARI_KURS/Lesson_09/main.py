from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import  system
wb = Workbook()

def product_bolim()->str:
    l_bolim = ['','Ichimliklar','Oziq_ovqat']
    for index,bolim in enumerate(l_bolim):
        if index != 0:
            print(index, bolim)
    i_bolim = input('Bo\'limni tanlang:')
    if i_bolim.isdigit() and int(i_bolim) < len(l_bolim) and int(i_bolim) > 0:
        i_bolim = int(i_bolim)
        return l_bolim[i_bolim] 
    elif i_bolim == '0':   
        main()
    else:
        print("Iltimos, bo'limga mos sonni tanlang!")
        main()
def product_name()->str:
    bolim = product_bolim()
    sheet = ['']
    data_bolim = load_workbook(f'database/{bolim}.xlsx')
    for sheet_name in data_bolim.sheetnames:
        sheet.append(sheet_name)
    for index,key in enumerate(sheet):
        if index != 0:
            print(index, key)
    select = input('Mahsulotni nomerini kiriting!/Yangi mahsulot nomini kiriting!/0:')
    if select == '0':
        main()
    elif select.isdigit() and int(select) < len(sheet) and int(select) > 0:
        select = int(select)
        return [bolim,sheet[select]]
    elif not (select.isdigit()):
        database = load_workbook(f'database/{bolim}.xlsx')
        new_sheet = database.create_sheet(select)
        database.save(f'database/{bolim}.xlsx')
        new_sheet.append(['Miqdori','Narxi','Chiqarilgan sanasi','Berilgan muddat','Kelgan sanasi','Chiqarilgan mamlakati'])
        return [bolim,select]
    else:
        print('Boshqatdan kiriting!')
        main()      

def product_name_sell()->str:
    bolim = product_bolim()
    sheet = ['']
    data_bolim = load_workbook(f'database/{bolim}.xlsx')
    for sheet_name in data_bolim.sheetnames:
        sheet.append(sheet_name)
    for index,key in enumerate(sheet):
        if index != 0:
            print(index, key)
    select = input('Mahsulotni nomerini kiriting!/0:')
    
    if select.isdigit() and int(select) < len(sheet) and int(select) > 0:
        select = int(select)
        return [bolim,sheet[select]] 
    else:
        print('Xatolik!')
        main() 
    
                 
def product_miqdori()->int:
    select = input('Mahsulot miqdorini kiriting!:')
    if select.isdigit():
        select = int(select)
        return select
    else:
        print('Iltimos son kiriting!')
        product_miqdori()
def product_price()->int:
    select = input('Mahsulot narxini kiriting!:')
    if select.isdigit():
        select = int(select)
        return select
    else:
        print('Iltimos son kiriting!')
        product_price()
        
def old_date()->datetime:
    select = input('Mahsulot chiqarilgan sanasini kiriting/  *kun/*oy/*yil:')
    if select.split('/')[2].isdigit() and int(select.split('/')[2]) > 1990 and int(select.split('/')[2]) <= 2024:
        year = int(select.split('/')[2])
        if select.split('/')[1].isdigit() and int(select.split('/')[1]) > 0 and int(select.split('/')[1]) <= 12:
            month = int(select.split('/')[1])
            if select.split('/')[0].isdigit() and int(month) > 0 and ((int(select.split('/')[0]) <= 31 and ((month % 2 == 1 and month <= 7) or (month % 2 == 0 and month > 7))) or (int(select.split('/')[0]) <= 30 and ((month % 2 == 0 and month <= 7 and month != 2) or (month % 2 == 1 and month > 7))) or (year % 4 != 0 and int(select.split('/')[0]) <= 28 and month == 2) or (year % 4 == 0 and int(select.split('/')[0]) <= 29 and month == 2)):
                day = int(select.split('/')[0])
                select = datetime(int(year),int(month),int(day))
                return select
            else:
                print('Kunni kiritishda xato!Iltimos qaytadan kiriting!')
                old_date()
        else:
            print('Oyni kiritishda xato!Iltimos qaytadan kiriting!')
            old_date()
    else:
        print('Yilni kiritishda xato!Iltimos qaytadan kiriting!')
        old_date()

def lifetime()->datetime:
    select = input('Mahsulot muddati tugaydigan sanasini kiriting/  *kun/*oy/*yil:')
    if select.split('/')[2].isdigit() and int(select.split('/')[2]) >= 2024 and int(select.split('/')[2]) <= 2050:
        year = int(select.split('/')[2])
        if select.split('/')[1].isdigit() and int(select.split('/')[1]) > 0 and int(select.split('/')[1]) <= 12:
            month = int(select.split('/')[1])
            if select.split('/')[0].isdigit() and int(select.split('/')[0]) > 0 and int(select.split('/')[0]) <= 31:
                day = int(select.split('/')[0])
                select = datetime(int(year),int(month),int(day))
                return select
            else:
                print('Kunni kiritishda xato!Iltimos qaytadan kiriting!')
                lifetime()
        else:
            print('Oyni kiritishda xato!Iltimos qaytadan kiriting!')
            lifetime()
    else:
        print('Yilni kiritishda xato!Iltimos qaytadan kiriting!')
        lifetime()     
def now_date()->datetime:
    select = datetime.now()
    return select
def country()->str:
    select = input('Mahsulot ishlab chiqarilgan mamlakatini kiriting:')
    return select

def add_data()->None:
    name = product_name().copy() 
    print(name)
    inf = [name[0],name[1],product_miqdori(),product_price(),old_date(),lifetime(),now_date(),country()].copy()
    add_sell = load_workbook('add-sell.xlsx')
    sheet = add_sell.active
    add_data = ['add-data',inf[1],inf[2],inf[3],inf[4],inf[5],inf[6],inf[7]]
    sheet.append(add_data)
    add_sell.save('add-sell.xlsx')
    
    data = load_workbook(f'database/{inf[0]}.xlsx') 
    add_database = [inf[2],inf[3],inf[4],inf[5],inf[6],inf[7]]
    for sheet_name in data.sheetnames:
        if sheet_name == inf[1]:
            sheet = data[sheet_name]
            sheet.append(add_database)
    data.save(f'database/{inf[0]}.xlsx')
    main()
def sell_data()->None:
    name = product_name_sell().copy()# name listi bolim va name ni qaytaradi
    data = load_workbook(f'database/{name[0]}.xlsx')
    sheet = data[name[1]]
    
    sell = load_workbook('add-sell.xlsx')
    sheet_sell = sell['Sheet1'] # sheet_sell bu sell faylini qaytaradi
    
    print(f'{name[1]} mahsulotimiz ushbu narxlarda:')
    price = []
    for i in range(2,sheet.max_row+1):
        price.append(sheet[f'B{i}'].value)
        print(i-1,'-->',sheet[f'B{i}'].value)
    select = input('Narxni tanlang:') 
    if select.isdigit() and int(select) <= len(price) and int(select) > 0:
        select = int(price[int(select)-1]) # select bu narxni belgilaydi
        print(select)
        miqdori = product_miqdori()
        jami = 0
        for i in range(2,sheet.max_row+1):
            if sheet[f'B{i}'].value == select:
                jami +=sheet[f'A{i}'].value
        if jami >= miqdori:
            for i in range(2,sheet.max_row+1):
                if sheet[f'B{i}'].value == select:
                    delta = sheet[f'A{i}'].value
                    sheet[f'A{i}'].value = sheet[f'A{i}'].value - miqdori  
                    data.save(f'database/{name[0]}.xlsx')
                    ilgargi_sheet = sheet[f'A{i}'].value
                    if sheet[f'A{i}'].value < 0:
                        sheet_sell.append(['sell-data',name[1],delta,select,sheet[f'C{i}'].value,sheet[f'D{i}'].value,datetime.now(),sheet[f'F{i}'].value])
                        sheet[f'A{i}'].value = 0
                        data.save(f'database/{name[0]}.xlsx')
                        sell.save('add-sell.xlsx')
                        miqdori = abs(ilgargi_sheet) 
                          
                    if sheet[f'A{i}'].value > 0:
                        sheet_sell.append(['sell-data',name[1],miqdori,select,sheet[f'C{i}'].value,sheet[f'D{i}'].value,datetime.now(),sheet[f'F{i}'].value])
                        sell.save('add-sell.xlsx')
                        miqdori = 0              
        else:
            print('Ushbu mahsulotimiz bu narxda yetarli miqdor mavjud emas!')
            sell_data()
        main()          
    else:
        print('Iltimos, narxni togri tanlang!')
        sell_data()
def databases()->None:
    select = input("Hisobot Bo'limni tanlang.1-qo'shish/2-sotish/3-database/1/2/3:")
    def add_sell()->list:
        add_sell = load_workbook('add-sell.xlsx')
        sheet = add_sell.active
        l_sheet = []
        for row in sheet.rows:
            i_sheet = []
            for element in row:
                i_sheet.append(element.value)
            l_sheet.append(i_sheet)
        del l_sheet[0]
        return l_sheet
    if select == '0':
        return main()
    elif select.strip() == '1':
        print('Yangi kelgan mahsulotlar:'.center(100))
        tartib = ['','Nomi','Miqdori','Narxi','Chiqarilgan sanasi','Muddati tugash sanasi','Kelgan vaqti','Mamlakat']
        for i in range(len(add_sell())):
            if add_sell()[i][0] == 'add-data':
                for j in range(len(add_sell()[i])):
                    if j != 0:
                        print(f'{tartib[j]} : {add_sell()[i][j]}')
            print()
        with open('add-data.txt', 'w', encoding='utf-8') as file:
            for i in range(len(add_sell())):
                if add_sell()[i][0] == 'add-data':
                    for j in range(len(add_sell()[i])):
                        if j != 0:
                            file.write(f'{tartib[j]} : {add_sell()[i][j]}\n')
                file.write('\n')
            file.close()
        system('add-data.txt')
        databases ()       
    elif select.strip() == '2':
        print('Sotilgan mahsulotlar:'.center(100))
        tartib = ['','Nomi','Miqdori','Narxi','Chiqarilgan sanasi','Muddati tugash sanasi','Kelgan vaqti','Mamlakat']
        for i in range(len(add_sell())):
            if add_sell()[i][0] == 'sell-data':
                for j in range(len(add_sell()[i])):
                    if j != 0:
                        print(f'{tartib[j]} : {add_sell()[i][j]}')
            print()
        with open('sell-data.txt', 'w', encoding='utf-8') as file:
            for i in range(len(add_sell())):
                if add_sell()[i][0] == 'sell-data':
                    for j in range(len(add_sell()[i])):
                        if j != 0:
                            file.write(f'{tartib[j]} : {add_sell()[i][j]}\n')
                file.write('\n')
            file.close()
        system('sell-data.txt')
        databases()
    elif select.strip() == '3':
        print('Ombordagi barcha mahsulotlar:'.center(150))
        database_ichimlik = load_workbook('database/Ichimliklar.xlsx')
        for sheet_name in database_ichimlik.sheetnames:
            sheet = database_ichimlik[sheet_name]
            print(sheet_name.center(150))
            for row in sheet.rows:
                for element in row:
                    print(str(element.value).center(20), end='| ')
                print()
            print() 
        database_oziq = load_workbook("database/Oziq_ovqat.xlsx")
        for sheet_name in database_oziq.sheetnames:
            sheet = database_oziq[sheet_name]
            print(sheet_name.center(150))
            for row in sheet.rows:    
                for element in row:
                    print(str(element.value).center(20), end='| ')
                print()
            print() 
        databases()
    else:
        print('Iltimos mos sonni tanlang!')
        databases()
def main()->None:
    select = input('Mahsulot qo\'shish,sotish va ma\'lumot olish uchun kerakli belgini mos ravishda tanlang/+/-/?:')
    if select == '+':
        add_data()
    elif select == '-':
        sell_data()
    elif select == '?':
        databases()
    elif select == '0':
        exit()
    else:
        print('Iltimos mos belgini tanlang!')
        main()
main()


        
        
    
    