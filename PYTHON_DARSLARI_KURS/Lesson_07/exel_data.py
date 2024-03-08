import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pickle

wb = Workbook()
ws = wb.active
workbook = load_workbook("MOCK_DATA1.xlsx")
sheet = workbook["data"]

list_question = ["",
                 "1)Ushbu jadvaldan oyligi 1000 $ dan yuqori bo'lgan ishchilarni saralab chiqaring",
                 "2)Jadvaldan ism yoki familiya bo'yicha turli insonlarni qidiring.Jadvalda mavjud bo'lsa,u haqidai barcha ma'lumotlarni chiroyli ko'rinishda qaytaring.Bo'lmasa,ushbu inson bizning ro'yxatda mavjud emas yozuvini chiqaring.",
                 "3)Hozirda 50 yoshdan katta shaxslarni saralang.Ularning barcha ma'lumotlarni consolega chiqaring.",
                 "4)Yangi sheet qo'shing.Bu sheetga 25 yoshli ishchilar ro'yxatini chiqaring.Ro'yxatda ularning barch ma'lumotlari aniq holda qaytarilsin.",
                 "5)Exelda berilgan ushbu ma'lumotlarni dictonary ko'rnishida matnli faylga o'tkazing.",
                 "6)Matnli fayldagi ma'lumotdan 30 yoshli shaxslarni ma'lumotlarini yangi exel faylga saqlang.",
                 "7)Jadvaldan shaxslarni first_name,last_name,salery si bo'yicha yangi sheetga nusxa ko'chiring.",
                 "8)Jurnaldagi shaxslarni oyligi bo'yicha diagramma qiling."
                ]
while True:
    for i in list_question:
        print(i, "\n")
    number = input("Savol nomeri:")

    if number.isdigit() and int(number) >0 and int(number) <=9:
        if int(number) == 1:
            list_inst = []
            for i in range(1000):
                string_salery = str(sheet[f'G{i + 2}'].value)
                string_salery = string_salery.replace('$', '')
                string_salery = float(string_salery)
                if string_salery > 1000:
                    string_salery = str(string_salery).replace('', '$', 1)
                    list_i = [sheet[f'A{i + 2}'].value, sheet[f'B{i + 2}'].value, sheet[f'C{i + 2}'].value, sheet[f'D{i + 2}'].value, sheet[f'E{i + 2}'].value, sheet[f'F{i + 2}'].value, string_salery]
                    list_inst.append(list_i)
            for i in list_inst:
                print()
                for j in i:
                    print(f'{str(j).center(25)}', end='  ')

        elif int(number) == 2:
            search_fl_name = input('Ism yoki familiya kiriting:')
            for i in range(ws.max_row-1):
                string_f_name = str(sheet[f'A{i + 2}'].value)
                string_l_name = str(sheet[f'B{i + 2}'].value)
                if (string_f_name or string_l_name) == search_fl_name:
                    print(str(string_f_name), "    ", str(string_l_name), "    ", str(sheet[f'C{i + 2}'].value), "    ", str(sheet[f'D{i + 2}'].value), "    ", str(sheet[f'E{i + 2}'].value), "    ", str(sheet[f'F{i + 2}'].value), "    ", str(sheet[f'G{i + 2}'].value))

        elif int(number) == 3:
            for i in range(ws.max_row-1):
                string_brith = sheet[f'F{i + 2}'].value
                print(string_brith)
                if isinstance(string_brith, str):
                    old_date = int(string_brith[-4:])
                    now_date = datetime.now()
                    yosh = now_date.year - old_date
                    if yosh > 50:
                        print(str(sheet[f'A{i + 2}'].value), "    ", str(sheet[f'B{i + 2}'].value), "    ", str(sheet[f'C{i + 2}'].value), "    ",str(sheet[f'D{i + 2}'].value), "    ", str(sheet[f'E{i + 2}'].value), "    ", str(sheet[f'F{i + 2}'].value), "    ",str(sheet[f'G{i + 2}'].value))
                else:
                    old_date = string_brith
                    now_date = datetime.now()

                    yosh = int(now_date.year) - int(old_date.year)
                    if yosh > 50:
                        print(str(sheet[f'A{i + 2}'].value), "    ", str(sheet[f'B{i + 2}'].value), "    ",str(sheet[f'C{i + 2}'].value), "    ", str(sheet[f'D{i + 2}'].value), "    ",str(sheet[f'E{i + 2}'].value), "    ", str(sheet[f'F{i + 2}'].value), "    ",str(sheet[f'G{i + 2}'].value))

        elif int(number) == 4:
            new_sheet = wb.create_sheet("25_yoshlilar")
            new_sheet.append([sheet['A1'].value, sheet['B1'].value, sheet['C1'].value, sheet['D1'].value, sheet['E1'].value, sheet['F1'].value, sheet['G1'].value])
            for i in range(ws.max_row-1):
                string_brith = sheet[f'F{i + 2}'].value
                if isinstance(string_brith, str):
                    old_date = int(string_brith[-4:])
                    now_date = datetime.now()
                    yosh = now_date.year - old_date
                    if yosh == 25:
                        new_sheet.append([sheet[f'A{i + 2}'].value, sheet[f'B{i + 2}'].value, sheet[f'C{i + 2}'].value, sheet[f'D{i + 2}'].value, sheet[f'E{i + 2}'].value, sheet[f'F{i + 2}'].value, sheet[f'G{i + 2}'].value])
                else:
                    old_date = string_brith
                    now_date = datetime.now()

                    yosh = int(now_date.year) - int(old_date.year)
                    if yosh == 25:
                        new_sheet.append([sheet[f'A{i + 2}'].value, sheet[f'B{i + 2}'].value, sheet[f'C{i + 2}'].value, sheet[f'D{i + 2}'].value, sheet[f'E{i + 2}'].value, sheet[f'F{i + 2}'].value, sheet[f'G{i + 2}'].value])
            wb.save("25_yoshlilar.xlsx")

        elif int(number) == 5:
            dict_data = {}
            list_inst = []
            for i in range(ws.max_row-1):
                list_i = [sheet[f'A{i + 2}'].value, sheet[f'B{i + 2}'].value, sheet[f'C{i + 2}'].value,sheet[f'D{i + 2}'].value, sheet[f'E{i + 2}'].value, sheet[f'F{i + 2}'].value, sheet[f'G{i + 2}'].value]
                dict_data.setdefault(f'id{i + 1}', list_i)
            with open('data', 'wb') as data:
                pickle.dump(dict_data, data)
            with open('data', 'rb') as data:
                dict_data = pickle.load(data)
            new_sheet2 = wb.create_sheet('matn')
            new_sheet2.append([sheet['A1'].value, sheet['B1'].value, sheet['C1'].value, sheet['D1'].value, sheet['E1'].value,sheet['F1'].value, sheet['G1'].value])
            for key, value in dict_data.items():
                list_i = []
                for v in value:
                    list_i.append(v)
                new_sheet2.append(list_i)

            wb.save("matn.xlsx")
        elif int(number) == 6:

            pass
        elif int(number) == 7:
            pass
        elif int(number) == 8:
            pass
        elif int(number) == 9:
            select = input("O'chirish fayl nomini to'liq kiriting:")
            if os.path.exists(select):
                os.rmdir(select)

            else:
                print("Bunday fayl yo'q!")
    else:
        print("Savol nomerini xato kiritdingiz!")
